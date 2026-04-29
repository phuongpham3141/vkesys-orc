"""File storage helpers for uploaded PDFs and exported results."""
from __future__ import annotations

import json
import uuid
from pathlib import Path
from typing import Iterable

from flask import current_app
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename


def upload_dir() -> Path:
    p = Path(current_app.config["UPLOAD_FOLDER"])
    p.mkdir(parents=True, exist_ok=True)
    return p


def output_dir() -> Path:
    p = Path(current_app.config["OUTPUT_FOLDER"])
    p.mkdir(parents=True, exist_ok=True)
    return p


def save_uploaded_pdf(file_storage: FileStorage) -> tuple[str, str, int]:
    """Persist an uploaded PDF to disk under a UUID name.

    Returns ``(stored_filename, original_filename, size_bytes)``.
    """
    original = secure_filename(file_storage.filename or "upload.pdf")
    if not original.lower().endswith(".pdf"):
        original = f"{original}.pdf"
    stored = f"{uuid.uuid4().hex}.pdf"
    path = upload_dir() / stored
    file_storage.save(str(path))
    size = path.stat().st_size
    return stored, original, size


def stored_path(stored_filename: str) -> Path:
    return upload_dir() / stored_filename


def remove_stored_file(stored_filename: str) -> None:
    try:
        stored_path(stored_filename).unlink(missing_ok=True)
    except OSError:
        pass


def export_results_text(job, results: Iterable) -> Path:
    """Write plain-text export of all pages."""
    target = output_dir() / f"job_{job.id}.txt"
    with target.open("w", encoding="utf-8") as fh:
        for r in results:
            fh.write(f"\n===== Trang {r.page_number} =====\n")
            fh.write(r.text_content or "")
            fh.write("\n")
    return target


def export_results_json(job, results: Iterable) -> Path:
    """Write structured JSON export."""
    target = output_dir() / f"job_{job.id}.json"
    payload = {
        "job_id": job.id,
        "filename": job.original_filename,
        "engine": job.engine,
        "page_count": job.page_count,
        "pages": [
            {
                "page_number": r.page_number,
                "text": r.text_content,
                "confidence": r.confidence_score,
            }
            for r in results
        ],
    }
    with target.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)
    return target


def export_results_markdown(job, results: Iterable) -> Path:
    target = output_dir() / f"job_{job.id}.md"
    with target.open("w", encoding="utf-8") as fh:
        fh.write(f"# {job.original_filename}\n\n")
        fh.write(f"- Engine: `{job.engine}`\n")
        fh.write(f"- Số trang: {job.page_count}\n\n")
        for r in results:
            fh.write(f"\n## Trang {r.page_number}\n\n")
            fh.write(r.text_content or "")
            fh.write("\n")
    return target


def _collect_tables(results: Iterable) -> list[tuple[int, int, list[list[str]]]]:
    """Pull every table out of results.

    Returns a list of ``(page_number, table_index_within_page, rows)`` tuples.
    Sources, in priority order:

      1. Structured cells stored in ``raw_response['tables']`` (Document AI).
      2. Markdown tables found in ``text_content`` (Gemini, Mistral).
    """
    out: list[tuple[int, int, list[list[str]]]] = []
    for r in results:
        raw = r.raw_response or {}
        structured = None
        if isinstance(raw, dict):
            structured = raw.get("tables")
        if structured and isinstance(structured, list):
            for tbl_idx, table in enumerate(structured):
                if isinstance(table, list) and table:
                    rows = [
                        [str(cell) for cell in row]
                        for row in table
                        if isinstance(row, list)
                    ]
                    if rows:
                        out.append((r.page_number, tbl_idx + 1, rows))
            continue
        # Fallback: parse Markdown tables from the page text.
        text = r.text_content or ""
        if "|" in text:
            for tbl_idx, table in enumerate(_parse_markdown_tables(text)):
                if table:
                    out.append((r.page_number, tbl_idx + 1, table))
    return out


def _parse_markdown_tables(text: str) -> list[list[list[str]]]:
    """Extract simple ``| a | b | c |`` Markdown tables from arbitrary text."""
    tables: list[list[list[str]]] = []
    current: list[list[str]] = []

    def flush() -> None:
        nonlocal current
        if current:
            tables.append(current)
        current = []

    sep_re = lambda s: bool(s) and all(  # noqa: E731
        c.strip("-: ") == "" for c in s.split("|") if c.strip()
    )

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if line.startswith("|") and line.endswith("|") and line.count("|") >= 2:
            inner = line[1:-1]
            if sep_re(inner):
                # delimiter row between header and body — ignore content
                continue
            cells = [c.strip() for c in inner.split("|")]
            current.append(cells)
        else:
            flush()
    flush()
    # Filter out 1-row "tables" that are probably noise
    return [t for t in tables if len(t) >= 2 and any(any(c for c in row) for row in t)]


def export_results_csv(job, results: Iterable) -> Path:
    """Export every detected table as CSV (UTF-8 BOM for Excel compatibility).

    Tables are concatenated, separated by ``=== Trang N – Bảng M ===`` rows.
    If no tables are found, falls back to a 2-column ``page,text`` dump so the
    download is never empty.
    """
    import csv

    results = list(results)
    target = output_dir() / f"job_{job.id}.csv"
    tables = _collect_tables(results)
    with target.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        if tables:
            for page_num, tbl_num, rows in tables:
                writer.writerow([f"=== Trang {page_num} – Bảng {tbl_num} ==="])
                for row in rows:
                    writer.writerow(row)
                writer.writerow([])
        else:
            writer.writerow(["page_number", "text"])
            for r in results:
                writer.writerow([r.page_number, r.text_content or ""])
    return target


def export_results_xlsx(job, results: Iterable) -> Path:
    """Export every detected table as a separate XLSX worksheet.

    Sheet naming: ``P{page}_T{table}`` (max 31 chars, Excel limit). Falls back
    to a single ``Pages`` sheet with ``page_number, text`` columns if no
    tables were detected.
    """
    from openpyxl import Workbook

    results = list(results)
    target = output_dir() / f"job_{job.id}.xlsx"
    wb = Workbook()
    # Workbook() always starts with a default sheet — remove it.
    if wb.worksheets:
        wb.remove(wb.active)

    tables = _collect_tables(results)
    if tables:
        for page_num, tbl_num, rows in tables:
            sheet_name = f"P{page_num}_T{tbl_num}"[:31]
            ws = wb.create_sheet(title=sheet_name)
            for row in rows:
                ws.append(row)
    else:
        ws = wb.create_sheet(title="Pages")
        ws.append(["Page", "Text"])
        for r in results:
            ws.append([r.page_number, r.text_content or ""])

    if not wb.worksheets:
        wb.create_sheet(title="Empty")
    wb.save(target)
    return target
